import os
import pandas as pd
from io import BytesIO

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils.timezone import now
from django.core.files.base import ContentFile, File
from django.core.files.storage import default_storage
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.conf import settings

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import FoodItem, Category, CartItem, Order, OrderHistory
from .forms import FoodItemForm, MenuUploadForm, CategoryForm
from .filters import FoodItemFilter
from django_filters.views import FilterView
from django.core.paginator import Paginator

@login_required
def home(request):
    categories = Category.objects.all()  # Fetch all categories
    return render(request, 'orders/home.html', {'categories': categories,'MEDIA_URL': settings.MEDIA_URL})

class CategoryMenuView(FilterView):
    model = FoodItem
    template_name = "orders/cat_menu.html"
    filterset_class = FoodItemFilter
    context_object_name = "food_items"
    paginate_by = 6

    def get_queryset(self):
        """Returns the filtered queryset."""
        print(self.request.GET)
        return self.filterset_class(self.request.GET, queryset=super().get_queryset()).qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Apply pagination correctly
        filtered_qs = self.get_queryset()
        paginator = Paginator(filtered_qs, self.paginate_by)
        page_number = self.request.GET.get("page")
        page_obj = paginator.get_page(page_number)

        context["page_obj"] = page_obj  # Standard name in Django CBVs
        context["food_items"] = page_obj  # So it works with the existing loop
        return context
    
    # def get_context_data(self, **kwargs):
    #     context = super().get_context_data(**kwargs)
        
    #     # Get filtered queryset
    #     filtered_qs = context["food_items"]

    #     # Paginate the queryset
    #     paginator = Paginator(filtered_qs, self.paginate_by)
    #     page = self.request.GET.get("page")
    #     food_items = paginator.get_page(page)
        
    #     context["food_items"] = food_items
    #     return context


# Cart
@login_required
def add_to_cart(request, food_id):
    if request.method == "POST":
        food_item = get_object_or_404(FoodItem, id=food_id)
        quantity = int(request.POST.get("quantity", 1))

        cart_item, created = CartItem.objects.get_or_create(user=request.user, food_item=food_item)

        if not created:
            cart_item.quantity += quantity  # Increment existing quantity
        else:
            cart_item.quantity = quantity  # Set new quantity

        cart_item.save()

        # Count total items in cart
        cart_count = CartItem.objects.filter(user=request.user).count()

        return JsonResponse({
            "message": f"Added {food_item.name} to cart!",
            "cart_count": cart_count
        })

    return JsonResponse({"error": "Invalid request"}, status=400)

@login_required
def cart_view(request):
    cart_items = CartItem.objects.filter(user=request.user)
    total_amount = sum(item.total_price() for item in cart_items)
    print("Cart Items:", cart_items)
    return render(request, 'orders/cart.html', {
        'cart_items': cart_items,
        'total_amount': total_amount
    })

@login_required
def remove_from_cart(request, item_id):
    cart_item = CartItem.objects.filter(user=request.user, id=item_id).first()

    if cart_item:
        cart_item.delete()
        messages.success(request, "Item removed from cart!")

    return redirect('orders:cart')

@login_required
def place_order(request):
    cart_items = CartItem.objects.filter(user=request.user)

    if not cart_items.exists():
        messages.error(request, "Your cart is empty.")
        return redirect('orders:cart')

    total_amount = sum(item.total_price() for item in cart_items)
    
    order = Order.objects.create(user=request.user, total_amount=total_amount,status= 'Completed')
    # order.items.set(cart_items)
    for cart_item in cart_items:
        OrderHistory.objects.create(
            order=order,
            item=cart_item.food_item,
            quantity=cart_item.quantity
        )


    cart_items.delete()  # Empty the cart

    messages.success(request, "Your order has been placed successfully!")
    return redirect('orders:payment_success')

@login_required
def payment_success(request):
    return render(request, 'orders/payment_success.html')

@login_required
def cancel_order(request):
    print("Cancel Order View Called")  # Debugging

    # Get the latest order of the user
    order = Order.objects.filter(user=request.user).order_by('-created_at').first()

    if not order:
        messages.error(request, "No recent order found to cancel.")
        return redirect('orders:home')

    # Calculate time difference
    time_elapsed = (now() - order.created_at).total_seconds()
    print(f"Time elapsed since order placement: {time_elapsed} seconds")  # Debugging

    if time_elapsed > 120:  # 120 seconds = 2 minutes
        messages.error(request, "Order cancellation time has expired.")
        return redirect('orders:home')

    # Debugging before canceling
    print(f"Canceling Order ID: {order.id}")

    order.status = 'Cancelled'
    order.save()
    # # Delete the order
    # order.delete()

    messages.success(request, "Your order has been Cancelled successfully.")
    return redirect('orders:home')

@login_required
def order_history(request):
    # Fetch all orders placed by the logged-in user
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    unique_statuses = set(orders.values_list('status', flat=True))
    print(unique_statuses)

    return render(request, 'orders/order_history_copy.html', {'orders': orders,'unique_statuses': unique_statuses,})

@staff_member_required  
def manage_menu(request):
    """Admin page to manage categories and food items"""
    if request.method == "POST":
        category_form = CategoryForm(request.POST)
        if category_form.is_valid():
            category_form.save()
            messages.success(request, "Category added successfully!")
            return redirect("orders:manage_menu")
        else:
            messages.error(request, "Error adding category. Please try again.")
    
    food_items = FoodItem.objects.all().order_by('category__name')
    categories = Category.objects.all()
    category_form = CategoryForm()

    return render(request, "orders/manage_menu.html", {
        "food_items": food_items,
        "categories": categories,
        "category_form": category_form
    })

@staff_member_required  
def delete_category(request, category_id):
    """Delete a category if it has no food items associated with it."""
    category = get_object_or_404(Category, id=category_id)
    
    if FoodItem.objects.filter(category=category).exists():
        messages.error(request, "Cannot delete category as it has food items associated with it.")
    else:
        category.delete()
        messages.success(request, "Category deleted successfully!")

    return redirect("orders:manage_menu")

@staff_member_required
def import_menu(request):
    if request.method == "POST":
        form = MenuUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES["file"]
            error_file_path = os.path.join(settings.MEDIA_ROOT, "menu_errors.xlsx")
            inserted_items = []
            has_errors = False
            error_log = []

            try:
                df = pd.read_excel(file)

                expected_columns = ["Name", "Category", "Price", "Description", "Is_Vegetarian", "Is_Vegan", "Image_Path"]

                # If file is empty
                if df.empty:
                    messages.error(request, "The uploaded file is empty.")
                    return redirect("orders:import_menu")

                # Identify missing or incorrect columns
                missing_columns = [col for col in expected_columns if col not in df.columns]
                extra_columns = [col for col in df.columns if col not in expected_columns]

                try:
                    wb = load_workbook(file)
                    ws = wb.active  # First sheet
                except:
                    wb = Workbook()
                    ws = wb.active
                    for col_num, col_name in enumerate(df.columns, 1):
                        ws.cell(row=1, column=col_num, value=col_name)

                error_highlight = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

                # Highlight incorrect column names
                for col_num, col_name in enumerate(df.columns, 1):
                    if col_name in missing_columns or col_name in extra_columns:
                        ws.cell(row=1, column=col_num).fill = error_highlight
                        has_errors = True

                if missing_columns or extra_columns:
                    messages.error(request, f"Incorrect file format. Missing columns: {', '.join(missing_columns)}")
                    messages.warning(request, f"Unexpected columns found: {', '.join(extra_columns)}")
                    wb.save(error_file_path)
                    return render(request, "orders/import_menu.html", {"form": form, "error_file": error_file_path})

                # Define Expected Data Types
                expected_dtypes = {
                    "Name": str,
                    "Category": str,
                    "Price": (int, float),
                    "Description": str,
                    "Is_Vegetarian": bool,
                    "Is_Vegan": bool,
                }

                # Validate row-wise data
                for index, row in df.iterrows():
                    row_num = index + 2  # Excel (1-based index + header row)
                    error_found = False
                    row_errors = []

                    for col, expected_type in expected_dtypes.items():
                        value = row.get(col, None)

                        # Check for missing values
                        if pd.isna(value):
                            row_errors.append(f"{col} is missing")
                            ws[f"{chr(65 + list(expected_dtypes.keys()).index(col))}{row_num}"].fill = error_highlight
                            error_found = True
                            continue

                        # Check data type
                        if not isinstance(value, expected_type):
                            expected_type_name = (
                                " or ".join(t.__name__ for t in expected_type)
                                if isinstance(expected_type, tuple)
                                else expected_type.__name__
                            )
                            row_errors.append(f"{col} must be {expected_type_name}")
                            ws[f"{chr(65 + list(expected_dtypes.keys()).index(col))}{row_num}"].fill = error_highlight
                            error_found = True

                    # Validate category
                    category, _ = Category.objects.get_or_create(name=row["Category"])

                    # Check for duplicate entries
                    exists = FoodItem.objects.filter(
                        name=row["Name"], category=category, is_vegetarian=row["Is_Vegetarian"], is_vegan=row["Is_Vegan"]
                    ).exists()

                    if exists:
                        row_errors.append(f"Duplicate entry: {row['Name']} in {row['Category']}")
                        ws[f"A{row_num}"].fill = error_highlight
                        error_found = True

                    if error_found:
                        error_log.append(f"Row {index+1}: {', '.join(row_errors)}")
                        has_errors = True
                        continue  # Skip inserting this row

                    # Insert valid food item
                    food_item = FoodItem(
                        name=row["Name"],
                        category=category,
                        price=row["Price"],
                        description=row["Description"],
                        is_vegetarian=row["Is_Vegetarian"],
                        is_vegan=row["Is_Vegan"],
                    )
                    food_item.save()
                    inserted_items.append(food_item.name)

                # Save error file if issues found
                if has_errors:
                    wb.save(error_file_path)
                    messages.error(request, "Errors found in your file. Download the error file for details.")
                    return render(request, "orders/import_menu.html", {"form": form, "error_file": error_file_path})

                if inserted_items:
                    messages.success(request, f"Inserted items: {', '.join(inserted_items)}")
                else:
                    messages.info(request, "No new food items were inserted.")

                return render(request, "orders/import_menu.html", {"form": form})

            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
                return redirect("orders:import_menu")

    else:
        form = MenuUploadForm()

    return render(request, "orders/import_menu.html", {"form": form})

# @staff_member_required 
# def import_menu(request):
#     if request.method == "POST":
#         form = MenuUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             file = request.FILES["file"]
#             error_file_path = os.path.join(settings.MEDIA_ROOT, "menu_errors.xlsx")
#             inserted_items = []
#             has_errors = False

#             try:
#                 df = pd.read_excel(file)
#                 expected_columns = ["Name", "Category", "Price", "Description", "Is_Vegetarian", "Is_Vegan", "Image_Path"]
                
#                 # If file is empty
#                 if df.empty:
#                     messages.error(request, "The uploaded file is empty.")
#                     return redirect("orders:import_menu")

#                 # Identify missing or incorrect columns
#                 missing_columns = [col for col in expected_columns if col not in df.columns]
#                 extra_columns = [col for col in df.columns if col not in expected_columns]
                
#                 try:
#                     wb = load_workbook(file)
#                     ws = wb.active  # First sheet
#                 except:
#                     wb = Workbook()
#                     ws = wb.active
#                     for col_num, col_name in enumerate(df.columns, 1):
#                         ws.cell(row=1, column=col_num, value=col_name)

#                 error_highlight = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                
#                 # Highlight incorrect column names
#                 for col_num, col_name in enumerate(df.columns, 1):
#                     if col_name in missing_columns or col_name in extra_columns:
#                         ws.cell(row=1, column=col_num).fill = error_highlight
#                         has_errors = True

#                 if missing_columns or extra_columns:
#                     messages.error(request, f"Incorrect file format. Missing columns: {', '.join(missing_columns)}")
#                     messages.warning(request, f"Unexpected columns found: {', '.join(extra_columns)}")
#                     wb.save(error_file_path)
#                     return render(request, "orders/import_menu.html", {"form": form, "error_file": error_file_path})
                
#                 # Validate row-wise data
#                 for index, row in df.iterrows():
#                     row_num = index + 2  # Adjust for Excel (1-based index + header row)
#                     error_found = False
                    
#                     # Validate category
#                     if pd.isna(row["Category"]) or not isinstance(row["Category"], str):
#                         ws[f"B{row_num}"].fill = error_highlight
#                         error_found = True
                    
#                     # Validate name
#                     if pd.isna(row["Name"]) or not isinstance(row["Name"], str):
#                         ws[f"A{row_num}"].fill = error_highlight
#                         error_found = True
                    
#                     # Validate price
#                     try:
#                         float(row["Price"])
#                     except (ValueError, TypeError):
#                         ws[f"C{row_num}"].fill = error_highlight
#                         error_found = True
                    
#                     # Validate boolean fields
#                     if row["Is_Vegetarian"] not in [True, False]:
#                         ws[f"E{row_num}"].fill = error_highlight
#                         error_found = True
#                     if row["Is_Vegan"] not in [True, False]:
#                         ws[f"F{row_num}"].fill = error_highlight
#                         error_found = True
                    
#                     # Check for duplicates
#                     category, _ = Category.objects.get_or_create(name=row["Category"])
#                     exists = FoodItem.objects.filter(
#                         name=row["Name"], category=category, is_vegetarian=row["Is_Vegetarian"], is_vegan=row["Is_Vegan"]
#                     ).exists()

#                     if exists:
#                         messages.warning(request, f"Skipping duplicate: {row['Name']} in {row['Category']}")
#                         ws[f"A{row_num}"].fill = error_highlight  # Highlight name as duplicate
#                         error_found = True
                    
#                     if error_found:
#                         has_errors = True
#                         continue  # Skip inserting this row
                    
#                     # Insert valid food item
#                     food_item = FoodItem(
#                         name=row["Name"],
#                         category=category,
#                         price=row["Price"],
#                         description=row["Description"],
#                         is_vegetarian=row["Is_Vegetarian"],
#                         is_vegan=row["Is_Vegan"],
#                     )
#                     food_item.save()
#                     inserted_items.append(food_item.name)
                
#                 # Save error file if issues found
#                 if has_errors:
#                     wb.save(error_file_path)
#                     messages.error(request, "Errors found in your file. Download the error file for details.")
                
#                 if inserted_items:
#                     messages.success(request, f"Inserted items: {', '.join(inserted_items)}")
#                 elif not has_errors:
#                     messages.info(request, "No new food items were inserted.")
                
#                 return render(request, "orders/import_menu.html", {"form": form, "error_file": error_file_path})
            
#             except Exception as e:
#                 messages.error(request, f"Error processing file: {str(e)}")
#                 return redirect("orders:import_menu")
    
#     else:
#         form = MenuUploadForm()
    
#     return render(request, "orders/import_menu.html", {"form": form})

# def import_menu(request):
#     if request.method == "POST":
#         form = MenuUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             file = request.FILES["file"]
#             inserted_items = []  # Track successfully inserted food items

#             try:
#                 df = pd.read_excel(file)  # Read the Excel file

#                 # Ensure expected columns exist
#                 # expected_columns = ["Name", "Category", "Price", "Description", "Is_Vegetarian", "Is_Vegan", "Image_Path"]
#                 # if not all(col in df.columns for col in expected_columns):
#                 #     messages.error(request, "Invalid file format. Ensure column names are correct.")
#                 #     return redirect("orders:import_menu")
                
#                 expected_columns = ["Name", "Category", "Price", "Description", "Is_Vegetarian", "Is_Vegan", "Image_Path"]
#                 missing_columns = [col for col in expected_columns if col not in df.columns]

#                 if missing_columns:
#                     print(missing_columns)
#                     messages.error(request, f"Invalid file format. Missing columns: {', '.join(missing_columns)}")
#                     return redirect("orders:import_menu")

#                 for _, row in df.iterrows():
#                     category, _ = Category.objects.get_or_create(name=row["Category"])  # Create or get category

#                     # Check if food item already exists (based on unique fields)
#                     print(f"Checking: {row['Name']} | {row['Category']} | {row['Is_Vegetarian']} | {row['Is_Vegan']}")
#                     exists = FoodItem.objects.filter(
#                         name=row["Name"],
#                         category=category,
#                         is_vegetarian=row["Is_Vegetarian"],
#                         is_vegan=row["Is_Vegan"],
#                     ).exists()

#                     # if exists:
#                     #     messages.warning(request, f"Skipping duplicate: {row['Name']} in {row['Category']}")
#                     #     continue  # Skip duplicate entry
#                     if exists:
#                         print(exists)
#                         messages.error(request, f"Skipping duplicate: {row['Name']} in {row['Category']}")  # Red alert
#                         continue 
#                     # Create food item
#                     food_item = FoodItem(
#                         name=row["Name"],
#                         category=category,
#                         price=row["Price"],
#                         description=row["Description"],
#                         is_vegetarian=row["Is_Vegetarian"],
#                         is_vegan=row["Is_Vegan"],
#                     )
#                     food_item.save()  # Save first to generate ID
#                     inserted_items.append(food_item.name)  # Add to inserted list

#                     # Process image if path is given
#                     image_path = str(row["Image_Path"]).strip() if row["Image_Path"] else ""
#                     if image_path:
#                         full_image_path = os.path.join(settings.MEDIA_ROOT, image_path)
#                         if os.path.exists(full_image_path):  # Check if file exists
#                             with open(full_image_path, "rb") as img_file:
#                                 ext = os.path.splitext(image_path)[-1]  # Get file extension
#                                 new_filename = f"{food_item.id}_{food_item.name}{ext}"  # Rename
#                                 image_file = File(img_file, name=new_filename)

#                                 # Save the image with new name
#                                 food_item.image.save(new_filename, image_file, save=True)

#                 if inserted_items:
#                     messages.success(request, f"Inserted items: {', '.join(inserted_items)}")
#                 else:
#                     messages.info(request, "No new food items were inserted.")

#                 return redirect("orders:import_menu")  # Redirect to Manage Menu

#             except Exception as e:
#                 messages.error(request, f"Error processing file: {str(e)}")
#                 return redirect("orders:import_menu")

#     else:
#         form = MenuUploadForm()

#     return render(request, "orders/import_menu.html", {"form": form})

# def import_menu(request):
#     if request.method == "POST":
#         form = MenuUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             file = request.FILES["file"]
#             try:
#                 df = pd.read_excel(file)  # Read the Excel file

#                 # Ensure expected columns exist
#                 expected_columns = ["Name", "Category", "Price", "Description", "Is_Vegetarian", "Is_Vegan", "Image_Path"]
#                 if not all(col in df.columns for col in expected_columns):
#                     messages.error(request, "Invalid file format. Ensure column names are correct.")
#                     return redirect("orders:import_menu")

#                 for _, row in df.iterrows():
#                     category, _ = Category.objects.get_or_create(name=row["Category"])  # Create or get category

#                     # Create food item (without image first, to get the ID)
#                     food_item = FoodItem(
#                         name=row["Name"],
#                         category=category,
#                         price=row["Price"],
#                         description=row["Description"],
#                         is_vegetarian=row["Is_Vegetarian"],
#                         is_vegan=row["Is_Vegan"],
#                     )
#                     food_item.save()  # Save first to generate ID

#                     # Process image if path is given
#                     # image_path = row["Image_Path"]
#                     image_path = str(row["Image_Path"]).strip() if row["Image_Path"] else ""
#                     if image_path:
#                         full_image_path = os.path.join(settings.MEDIA_ROOT, image_path)
#                         if os.path.exists(full_image_path):  # Check if file exists
#                             with open(full_image_path, "rb") as img_file:
#                                 ext = os.path.splitext(image_path)[-1]  # Get file extension
#                                 new_filename = f"{food_item.id}_{food_item.name}{ext}"  # Rename
#                                 image_file = File(img_file, name=new_filename)

#                                 # Save the image with new name
#                                 food_item.image.save(new_filename, image_file, save=True)

#                 messages.success(request, "Menu items imported successfully!")
#                 return redirect("orders:manage_menu")  # Redirect to Manage Menu

#             except Exception as e:
#                 messages.error(request, f"Error processing file: {str(e)}")
#                 return redirect("orders:import_menu")

#     else:
#         form = MenuUploadForm()

#     return render(request, "orders/import_menu.html", {"form": form})

@staff_member_required
def add_food_item(request):
    if request.method == "POST":
        form = FoodItemForm(request.POST, request.FILES)
        if form.is_valid():
            food_item = form.save(commit=False)  # Don't save to DB yet

            if food_item.image and food_item.image.name != "food_item_pics/food_default.png":
                ext = food_item.image.name.split('.')[-1]  # Get file extension
                food_item.save()  # Save to assign an ID

                # Generate new filename: id_name.ext
                new_filename = f"food_item_pics/{food_item.id}_{food_item.name}.{ext}"

                # Save the file with the new name
                old_path = food_item.image.path  # Old file path
                new_path = os.path.join(default_storage.location, new_filename)

                # Rename file if old_path exists
                if os.path.exists(old_path):
                    os.rename(old_path, new_path)

                # Update model with new file name
                food_item.image.name = new_filename
                food_item.save(update_fields=['image'])  # Save again to update image field
            else:
                food_item.image = "food_item_pics/food_default.png"  # Ensure default image stays the same
                food_item.save()

            return redirect('orders:manage_menu')  # Redirect to menu management page
    else:
        form = FoodItemForm()
    
    return render(request, "orders/add_food_item.html", {"form": form})

@staff_member_required
def update_food_item(request, food_id):
    """Admin can update existing food items"""
    food_item = get_object_or_404(FoodItem, id=food_id)
    if request.method == "POST":
        form = FoodItemForm(request.POST, request.FILES, instance=food_item)
        if form.is_valid():
            form.save()
            return redirect('orders:manage_menu')
    else:
        form = FoodItemForm(instance=food_item)
    return render(request, 'orders/update_food_item.html', {'form': form, 'food_item': food_item})

@staff_member_required
def delete_food_item(request, food_id):
    """Admin can delete food items"""
    food_item = get_object_or_404(FoodItem, id=food_id)
    if request.method == "POST":
        food_item.delete()
        return redirect('orders:manage_menu')
    return render(request, 'orders/delete_food_item.html', {'food_item': food_item})

